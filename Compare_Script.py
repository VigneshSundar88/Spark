import ConfigParser as CP
import subprocess as SP
import commands
import os
import time
import sys
import pandas as pd
import openpyxl as op
from collections import defaultdict

from pyspark import SparkContext
from pyspark.sql import SparkSession, DataFrame, Window
from pyspark.sql import functions as func
from pyspark.sql.types import StringType

start = time.time()

class comparison_script():
	print("Script Launched")
	def __init__(self):
	    self.config_parse = CP.ConfigParser()
		self.beeline_connection = ""
		self.source = None
		self.target = None
		self.source_name_list = []
		self.target_name_list = []
		self.source_data = None
		self.target_data = None
		self.source_name = None
		self.target_name = None
		self.source_df = None
		self.target_df = None
		self.sep = None
		self.read_source_changed_df = None
		self.read_target_changed_df = None
		self.source_df_list = []
		self.target_df_list = []
		self.source_key_list = []
		self.target_key_list = []
		self.source_changed_key = None
		self.target_changed_key = None
		self.on_cond = None
		self.output_summ_file_location_list = []
		self.output_summ_file_name_list = []
		
	def assign_self_values(self):
		open_config_file = open(os.path.abspath(".")+'/comparison_params.ini')
		self.config_parse.readfp(open_config_file)
		self.source = self.config_parse.get('COMPARISON_FORMAT', 'SOURCE')
		self.target = self.config_parse.get('COMPARISON_FORMAT', 'TARGET')
		self.source_name_list = self.config_parse.items('SOURCE_NAME')
		self.target_name_list = self.config_parse.items('TARGET_NAME')
		self.source_data = self.config_parse.items('SOURCE')
		self.target_data = self.config_parse.items('TARGET')
		self.sep = self.config_parse.items('SEP')
		self.output_summ_file_location_list = self.config_parse.items('OUTPUT_FILE_LOCATION')
		self.output_summ_file_name_list = self.config_parse.items('OUTPUT_FILE_NAME')

	def create_source_target_df(self):
		if self.source.lower() == "hive" or self.target.lower() == "hive":
			if self.source.lower() == "hive":
				tbl_list = self.source_data
			elif self.target.lower() == "hive":
				tbl_list = self.target_data
				
			if len(tbl_list) > 1:
				for tbl_pos, hive_tbl in enumerate(tbl_list):
					hive_tbl_name = hive_tbl[1]
					
					if self.source.lower() == "hive":
						self.source_df = spark.read.table(hive_tbl_name)
						self.source_df_list.append(self.source_df)
					elif self.target.lower() == "hive":
						self.target_df = spark.read.table(hive_tbl_name)
						self.target_df_list.append(self.target_df)
			elif len(tbl_list) == 1:
				hive_tbl_name = tbl_list[0][1]
				if self.source.lower() == "hive":
					self.source_df = spark.read.table(hive_tbl_name)
					self.source_df_list.append(self.source_df)
				elif self.target.lower() == "hive":
					self.target_df = spark.read.table(hive_tbl_name)
					self.target_df_list.append(self.target_df)
					
			if len(csv_list) > 1:
				for csv_pos, csv_file in enumerate(csv_list):
					csv_file_name = csv_file[1]
					
					if self.source.lower() == "csv":
					    sep_val = self.sep[csv_pos][1]
						self.source_df = spark.read.csv(csv_file_name, sep=sep_val, header=True)
						self.source_df_list.append(self.source_df)
					elif self.target.lower() == "csv":
						sep_val = self.sep[csv_pos][1]
						self.target_df = spark.read.csv(csv_file_name, sep=sep_val, header=True)
						self.target_df_list.append(self.target_df)
			elif len(csv_list) == 1:
				sep_val = self.sep[0][1]
				csv_file_name = csv_list[0][1]
				if self.source.lower() == "csv":
					self.source_df = spark.read.csv(csv_file_name, sep=sep_val, header=True)
					self.source_df_list.append(self.source_df)
				elif self.target.lower() == "hive":
					self.target_df = spark.read.csv(csv_file_name, sep=sep_val, header=True)
					self.target_df_list.append(self.target_df)
					
	def change_df_col(self):
		for ind in range(len(self.source_df_list)):
			source_df = self.source_df_list[ind]
			target_df = self.target_df_list[ind]
			
			source_df_init_columns = source_df.columns
			target_df_init_columns = target_df.columns
			
			source_col_name_lower = map(lambda col_name: col_name.lower(), source_df_init_columns)
			target_col_name_lower = map(lambda col_name: col_name.lower(), target_df_init_columns)
			
			common_columns = list(set(source_col_name_lower).intersection(set(target_col_name_lower)))
			
			read_source_df = source_df.select(common_columns)
			read_target_df = target_df.select(common_columns)
			
			read_source_df_columns = map(lambda col_name: "source_"+col_name, common_columns)
			self.read_source_changed_df = read_source_df.toDF(*read_source_df_columns)
			self.read_source_changed_df.cache()
			
			read_target_df_columns = map(lambda col_name: "target_"+col_name, common_columns)
			self.read_target_changed_df = read_target_df_columns.toDF(*read_target_df_columns)
			self.read_target_changed_df.cache()
			
			self.create_key_columns(pos_val=ind)
			self.source_name = self.source_name_list[ind][1]
			self.target_name = self.target_name_list[ind][1]
			
			print("Comparison between the {source_name} and {target_name}".format(source_name=self.source_name, target_name=self.target_name))
			
			map(lambda col_name: self.compare_source_and_target_df(col_name, pos_val=ind), common_columns)
			
	def create_key_columns(self, pos_val):
		fetch_all_source_key = self.config_parse.items('SOURCE_KEYS')
		fetch_exact_source_key = fetch_all_source_key[ind][1]
		self.source_key_list = fetch_exact_source_key.split(",")
		print("source_key_list", self.source_key_list)
		source_key_list = ["source_"+k.lower().strip() for k in self.source_key_list]
		
		fetch_all_target_key = self.config_parse.items('TARGET_KEYS')
		fetch_exact_target_key = fetch_all_target_key[ind][1]
		self.target_key_list = fetch_exact_target_key.split(",")
		print("target_key_list", self.target_key_list)
		target_key_list = ["target_"+k.lower().strip() for k in self.target_key_list]
		
		self.source_changed_key = source_key_list
		self.target_changed_key = target_key_list

		self.on_cond = [func.col(f) == func.col(s) for (f, s) in zip(self.source_changed_key, self.target_changed_key)]
		
	def compare_source_and_target_df(self, tbl_col_name, pos_val):
		diff_val_df = defaultdict(list)
		diff_summary_df = {}
		
		print("Getting the {col_name_val} data ready for comparison...".format(col_name_val = tbl_col_name))
		
		source_changed_col_name = "source_"+tbl_col_name
		target_changed_col_name = "target_"+tbl_col_name
		
		if source_changed_col_name.lower() not in self.source_changed_key:
			source_select_list = self.source_changed_key[:]
			source_select_list.append(source_changed_col_name)
			source_join_col_df = self.read_source_changed_df.select(source_select_list).sortWithinPartitions(self.source_changed_key)
			source_ind_col_count = source_join_col_df.select(source_changed_col_name).count()
			
			target_select_list = self.target_changed_key[:]
			target_select_list.append(target_changed_col_name)
			target_join_col_df = self.read_target_changed_df.select(target_select_list).sortWithinPartitions(self.target_changed_key)
			target_ind_col_count = target_join_col_df.select(target_changed_col_name).count()
			
			source_target_join_outer = source_join_col_df.join(target_join_col_df, self.on_cond, 'outer')\
									   .withColumn("matches", func.when(func.trim(func.col("{source_col_name_val}".format(source_col_name_val=source_changed_col_name))) == func.trim(func.col("{target_col_name_val}".format(target_col_name_val=target_changed_col_name))), 1).otherwise(0))\
									   .withColumn("mismatches", func.when(func.trim(func.col("{source_col_name_val}".format(source_changed_col_name=source_changed_col_name))) != func.trim(func.col("{target_col_name_val}".format(target_col_name_val=target_changed_col_name))), 1).otherwise(0))\
									   .withColumn("missing_in_src", func.when(func.trim(func.col("{source_col_name_val}".format(source_col_name_val=source_changed_col_name))).isNull() & func.trim(func.col("{target_col_name_val}".format(target_col_name_val=target_changed_col_name))).isNotNull(), 1).otherwise(0))\
									   .withColumn("missing_in_tgt", func.when(func.trim(func.col("{source_col_name_val}".format(source_col_name_val=source_changed_col_name))).isNotNull() & func.trim("{target_col_name_val}".format(target_col_name_val=target_changed_col_name)).isNull(), 1).otherwise(0))
								
			mismatch_val_count_df = source_target_join_outer.select(func.lit("BLUESTAR").alias("SOURCE"), func.lit(self.source_name).alias("source_name"), func.lit(self.target_name).alias("target_name"), func.lit(tbl_col_name).alias("COLUMN_NAME"), "*")
			
			mismatch_val_full_df = mismatch_val_count_df.where("mismatches = 1")
			mismatch_val_limit_df = self.filter_df_count(mismatch_val_full_df)
			self.diff_val_df["mismatch_val"]["mismatch_val_{raw_col_name}".format(raw_col_name=tbl_col_name)] = mismatch_val_limit_df
			
			missing_in_src_full_df = mismatch_val_count_df.where("missing_in_src = 1")
			missing_in_src_limit_df = self.filter_df_count(missing_in_src_full_df)
			self.diff_val_df["missing_in_src"]["missing_in_src_{raw_col_name}".format(raw_col_name=tbl_col_name)] = missing_in_src_limit_df
			
			missing_in_tgt_full_df = mismatch_val_count_df.where("missing_in_tgt = 1")
			missing_in_tgt_limit_df = self.filter_df_count(missing_in_tgt_full_df)
			self.diff_val_df["missing_in_tgt"]["missing_in_tgt_{raw_col_name}".format(raw_col_name=tbl_col_name)] = missing_in_tgt_limit_df
			
			summ_df_sum = source_target_join_outer.select(func.col("matches"), func.col("mismatches"), func.col("missing_in_src"), func.col("missing_in_tgt")).groupBy().sum()
			
			summ_df_sum_alias = summ_df_sum.withColumnRenamed("sum(matches)", "matches").withColumnRenamed("sum(mismatches)", "mismatches").withColumnRenamed("sum(missing_in_src)", "missing_in_src").withColumnRenamed("sum(missing_in_tgt)", "missing_in_tgt")
			
			summ_df_count = summ_df_sum_alias.withColumn("source_count", func.lit(source_ind_col_count)).withColumn("target_count", func.lit(target_ind_col_count)).withColumn("SOURCE", func.lit("BLUESTAR")).withColumn("source_name", func.lit(self.source_name)).withColumn("target_name", func.lit(self.target_name)).withColumn("COLUMN_NAME", func.lit(tbl_col_name)).select("SOURCE", "SOURCE_NAME", "TARGET_NAME", "COLUMN_NAME", "source_count", "target_count", "matches", "mismatches", "missing_in_src", "missing_in_tgt")
			
			self.diff_summary_df_created["diff_summ_df_{col_name}".format(col_name = tbl_col_name)] = summ_df_count
			
	def filter_df_count(self, ind_full_df):
	    ind_full_df_count = ind_full_df.count()
		if ind_full_df_count <= 50 and ind_full_df_count != 0:
		    ind_limit_df = ind_full_df.limit(ind_full_df_count)
		elif ind_full_df_count > 50:
		    ind_limit_df = ind_full_df.limit(10)
		elif ind_full_df_count == 0:
		    ind_limit_df = ind_full_df
		return ind_limit_df
		
	def union_tbl_summ_diff(self, position_val):
	    print("Consolidation of comparison summary is in progress...")
		
		output_file_location = self.output_summ_file_location_list[position_val][1]
		output_file_name = self.output_summ_file_name_list[position_val][1]
		
		wb = op.Workbook()
		worksheet_names = self.diff_val_df.keys()
		map(lambda ws_name: wb.create_sheet(ws_name), worksheet_names)
		fetch_sheet = wb["Sheet"]
		fetch_sheet.title = "Comparison_Summary"
		wb.save('{output_file_location}/{output_file_name}'.format(output_file_location = output_file_location, output_file_name = output_file_name))
		#print(wb.sheetnames)
		
		summary_tbl_col = self.diff_summary_df_created.values()[0].columns
		summary_col_name = self.source_changed_key+self.target_changed_key+summary_tbl_col
		
		union_summ_col_df = reduce(lambda df1, df2: df1.union(df2), self.diff_summary_df_created.values())
		number_of_columns = union_summ_col_df.count()
		union_summ_col_df.show(number_of_columns, truncate=False)
		self.write_to_output_file(worksheet_name="Comparison_Summary", df_val=union_summ_col_df, pos_val=position_val)
		
		print("Consolidation of mismatch_values is in progress...")
		
		print("self.diff_val_df_keys", self.diff_val_df.keys())
		
		for summ_tbl_name, summ_col_name in self.diff_val_df.items():
		    print(summ_tbl_name)
			col_summary_dict = {"mismatch_val": "mismatching", "missing_in_src": "missing in source", "missing_in_tgt", "missing in target"}
			no_val_col_list = []
		
			for df_val in summ_col_name.values():
			    df_val_count = df_val.count()
				unique_col = list(set(df_val.columns) - set(summary_col_name))
                no_val_dup_col_name = map(lambda col_val: col_val.split("_", 1)[1], unique_col)
				no_val_uni_col_name = list(set(no_val_dup_col_name))[0]
				
				if df_val_count:
				    comparison_col_name = no_val_uni_col_name.upper()+" :"
					print(comparison_col_name)
					final_df_summ = df_val.drop(*['matches', 'mismatches', 'missing_in_src', 'missing_in_tgt']).limit(df_val_count)
                    self.write_to_output_file(worksheet_name=summ_tbl_name, tbl_heading=comparison_col_name, df_val=final_df_summ, pos_val=position_val)
                else:
                    no_val_col_list.append(no_val_uni_col_name)

            no_record_cols = "There is no {summary_col_name} record for the column(s): {no_val_col}\n".format(summary_col_name = col_summary_dict[summ_tbl_name], no_val_col_val=no_val_col_list)
            print(no_record_cols)
			self.write_to_output_file(worksheet_name=summ_tbl_name, no_val_col=no_record_cols, pos_val=position_val)
			self.diff_summary_df_created.clear()
			self.diff_val_df.clear()
			
	def write_to_output_file(self, worksheet_name=None, tbl_heading=None, df_val=None, no_val_col=None, pos_val=None):
	    output_file_location = self.output_summ_file_location_list[pos_val][1]
		output_file_name = self.output_summ_file_name_list[pos_val][1]
		
		writer = pd.ExcelWriter('{file_location}/{file_name}'.format(file_location=output_file_location, file_name=output_file_name), mode='a', engine='openpyxl')
		book = op.load_workbook('{file_location}/{file_name}'.format(file_location=output_file_location, file_name=output_file_name))
		writer.book = book
		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
		
		if df_val != None:
		    conv_spark_df_pd_df = df_val.toPandas()
			conv_spark_df_pd_df_str = conv_spark_df_pd_df.astype(str)
			start_row = writer.sheets[worksheet_name].max_row + 2
			book_read_sheet = book.get_sheet_by_name(worksheet_name)
			book_read_sheet['A{cell_val}'.format(cell_val=start_row)] = tbl_heading
			conv_spark_df_pd_df_str.to_excel(writer, startrow=start_row, sheet_name=worksheet_name, na_rep='Null', index=False)
		elif df_val == None and no_val_col != None:
		    start_row = writer.sheets[worksheet_name].max_row + 2
			book_read_sheet = book.get_sheet_by_name(worksheet_name)
			book_read_sheet['A{cell_val}'.format(cell_val=start_row)] = no_val_col
		writer.save()
		
if __name__ == "__main__":
    spark = SparkSession\
	        .builder\
			.enableHiveSupport()\
			.getOrCreate()
			
    compare_script_obj = comparison_script()
	compare_script_obj.assign_self_values()
	compare_script_obj.create_source_target_df()
	compare_script_obj.change_df_col
	
end = time.time()
difference = int(end-start)
print(str(difference)+ " secs")